#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Génère un fichier HTML unique pour les 4 corpus XLSX.

Interface restructurée en deux sections :
  - « Affichage » : contrôle visuel (surlignages, badges, élongations)
  - « Filtrer »   : sélection des données (corpus, émotions, traits binaires,
                     variables qualitatives, recherche texte)

Les colonnes binaires (ironie, insulte, mépris / haine, argot, abréviation) et
qualitatives (ROLE, HATE, SENTIMENT, TARGET, VERBAL_ABUSE, INTENTION, CONTEXT)
sont extraites des XLSX et injectées comme data-* attributes pour un filtrage
réactif côté client.
"""

from __future__ import annotations

import html
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

from _viz_template import CSS, JAVASCRIPT

# ── Constantes ────────────────────────────────────────────────────────────────

MODE_COLORS = {
    "Désignée": "rgba(40, 80, 160, 0.30)",
    "Comportementale": "rgba(0, 130, 100, 0.30)",
    "Suggérée": "rgba(180, 100, 20, 0.30)",
    "Montrée": "rgba(120, 50, 140, 0.30)",
}
MODE_COLORS_SOLID = {
    "Désignée": "rgb(40, 80, 160)",
    "Comportementale": "rgb(0, 130, 100)",
    "Suggérée": "rgb(180, 100, 20)",
    "Montrée": "rgb(120, 50, 140)",
}
MODES = list(MODE_COLORS.keys())

EMOTION_LABELS = [
    "Colère", "Dégoût", "Joie", "Peur", "Surprise", "Tristesse",
    "Admiration", "Culpabilité", "Embarras", "Fierté", "Jalousie", "Autre",
]

COL_RENAME = {
    "Designee": "Désignée", "Montree": "Montrée", "Suggeree": "Suggérée",
    "Colere": "Colère", "Degout": "Dégoût",
    "Culpabilite": "Culpabilité", "Fierte": "Fierté",
}

BINARY_FEATURES = [
    {"col": "ironie",          "attr": "ironie",      "label": "Ironie"},
    {"col": "insulte",         "attr": "insulte",     "label": "Insulte"},
    {"col": "mépris / haine",  "attr": "mepris",      "label": "Mépris / Haine"},
    {"col": "argot",           "attr": "argot",        "label": "Argot"},
    {"col": "abréviation",     "attr": "abreviation",  "label": "Abréviation"},
]

QUAL_FEATURES = [
    {"col": "ROLE",          "attr": "role",        "label": "ROLE"},
    {"col": "HATE",          "attr": "hate",        "label": "HATE"},
    {"col": "SENTIMENT",     "attr": "sentiment",   "label": "SENTIMENT"},
    {"col": "TARGET",        "attr": "target",      "label": "TARGET"},
    {"col": "VERBAL_ABUSE",  "attr": "verbalabuse", "label": "VERBAL_ABUSE"},
    {"col": "INTENTION",     "attr": "intention",   "label": "INTENTION"},
    {"col": "CONTEXT",       "attr": "context",     "label": "CONTEXT"},
]

DATASETS = [
    {"key": "homophobie", "label": "Homophobie", "path": "outputs/homophobie_annotations_gold_flat_updated.xlsx"},
    {"key": "obesite",    "label": "Obésité",    "path": "outputs/obésité_annotations_gold_flat_updated.xlsx"},
    {"key": "racisme",    "label": "Racisme",    "path": "outputs/racisme_annotations_gold_flat_updated.xlsx"},
    {"key": "religion",   "label": "Religion",   "path": "outputs/religion_annotations_gold_flat_updated.xlsx"},
]

DESACCORD_LABEL = "Désaccord"
BASE_DIR = Path("/home/gwen/Annotations_v1")
OUTPUT_PATH = BASE_DIR / "outputs/viz_all_corpora_unified_v2.html"

# ── Fonctions utilitaires ─────────────────────────────────────────────────────

def clean_qualitative_value(val) -> str:
    """Nettoie une valeur qualitative : strip, détection Majority: NULL → Désaccord."""
    if val is None or str(val).strip() == "":
        return ""
    s = str(val).strip()
    if "Majority: NULL" in s:
        return DESACCORD_LABEL
    return s


def normalise_record(row: dict) -> dict:
    out = {COL_RENAME.get(k, k): v for k, v in row.items()}
    if "spans_json" not in out and "_span_details" in out:
        out["spans_json"] = out["_span_details"]
    if "text" not in out and "TEXT" in out:
        out["text"] = out["TEXT"]
    # Nettoyer les qualitatives
    for qf in QUAL_FEATURES:
        col = qf["col"]
        if col in out:
            out[col] = clean_qualitative_value(out[col])
    # Nettoyer les binaires → int
    for bf in BINARY_FEATURES:
        out[bf["col"]] = as_int(out.get(bf["col"]))
    return out


def read_xlsx(path: Path) -> tuple[list[str], list[tuple]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]
    wb.close()
    return headers, rows


def rows_to_dicts(headers: list[str], rows: list[tuple]) -> list[dict]:
    return [normalise_record(dict(zip(headers, row))) for row in rows]


def as_int(value) -> int:
    if value in (None, "", False):
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def get_active_modes(row: dict) -> list[str]:
    return [mode for mode in MODES if as_int(row.get(mode)) == 1]


def get_active_emotions(row: dict) -> list[str]:
    return [emo for emo in EMOTION_LABELS if as_int(row.get(emo)) == 1]


def collect_qual_values(all_records: list[dict]) -> dict[str, list[str]]:
    """Collecte les valeurs uniques par colonne qualitative, en séparant les /."""
    vals: dict[str, set[str]] = {qf["col"]: set() for qf in QUAL_FEATURES}
    for row in all_records:
        for qf in QUAL_FEATURES:
            v = row.get(qf["col"], "")
            if v:
                # Séparer les valeurs composites (ex: "bully/bully_support")
                for part in str(v).split("/"):
                    p = part.strip()
                    if p:
                        vals[qf["col"]].add(p)
    return {col: sorted(vs) for col, vs in vals.items()}


# ── Statistiques ──────────────────────────────────────────────────────────────

def compute_stats(records: list[dict]) -> dict:
    mode_counts = Counter()
    emotion_counts = Counter()
    for row in records:
        mode_counts.update(get_active_modes(row))
        emotion_counts.update(get_active_emotions(row))
    return {
        "total": len(records),
        "mode_counts": {mode: mode_counts.get(mode, 0) for mode in MODES},
        "emo_counts": {emo: emotion_counts.get(emo, 0) for emo in EMOTION_LABELS if emotion_counts.get(emo, 0)},
    }


# ── Rendu du texte avec surlignages ──────────────────────────────────────────

def highlight_text_with_spans(text: str, spans_json_str: str | None,
                              elongation_spans: list[tuple[int, int]] | None = None) -> str:
    """Rendu du texte avec surlignages émotionnels et élongations.

    Chaque span reçoit un attribut data-modes pour le contrôle d'affichage JS.
    """
    if not text:
        return ""
    has_emotion = spans_json_str and spans_json_str != "[]"
    has_elong = bool(elongation_spans)
    if not has_emotion and not has_elong:
        return html.escape(text)

    char_modes: list[set] = [set() for _ in range(len(text))]
    if has_emotion:
        try:
            spans = json.loads(spans_json_str)
        except (json.JSONDecodeError, TypeError):
            spans = []
        text_lower = text.lower()
        for span in spans:
            span_text = (span or {}).get("span_text", "")
            mode = (span or {}).get("mode", "")
            if not span_text or mode not in MODE_COLORS:
                continue
            idx = text_lower.find(str(span_text).lower())
            if idx >= 0:
                for pos in range(idx, min(idx + len(span_text), len(text))):
                    char_modes[pos].add(mode)

    char_elong: list[bool] = [False] * len(text)
    if has_elong:
        for start, end in elongation_spans:
            for pos in range(start, min(end, len(text))):
                char_elong[pos] = True

    parts: list[str] = []
    i = 0
    while i < len(text):
        modes = frozenset(char_modes[i])
        elong = char_elong[i]
        j = i + 1
        while j < len(text) and frozenset(char_modes[j]) == modes and char_elong[j] == elong:
            j += 1
        chunk = html.escape(text[i:j])
        style_parts: list[str] = []
        classes: list[str] = []
        title_parts: list[str] = []
        data_attrs = ""

        if modes:
            sorted_modes = sorted(modes)
            style_parts.append(f"background:{MODE_COLORS[sorted_modes[0]]}")
            if len(sorted_modes) > 1:
                style_parts.append(f"border-bottom:2px solid {MODE_COLORS_SOLID[sorted_modes[1]]}")
            title_parts.extend(sorted_modes)
            classes.append("hl")
            data_attrs = f' data-modes="{html.escape(",".join(sorted_modes), quote=True)}"'

        if elong:
            classes.append("elong")
            title_parts.append("Élongation")

        if classes:
            cls_attr = " ".join(classes)
            style_attr = ";".join(style_parts)
            title_attr = html.escape(" + ".join(title_parts), quote=True)
            parts.append(f'<span class="{cls_attr}" style="{style_attr}" title="{title_attr}"{data_attrs}>{chunk}</span>')
        else:
            parts.append(chunk)
        i = j
    return "".join(parts)


# ── Rendu HTML ────────────────────────────────────────────────────────────────

def bar_chart_html(counts: dict, color_map: dict | None = None, max_width: int = 220) -> str:
    if not counts:
        return ""
    max_val = max(counts.values(), default=0)
    if max_val == 0:
        return ""
    parts = []
    for label, value in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
        width = int(value / max_val * max_width) if max_val else 0
        color = (color_map or {}).get(label, "#6366f1")
        parts.append(
            f'<div class="bar-row">'
            f'<span class="bar-label">{html.escape(label)}</span>'
            f'<div class="bar" style="width:{width}px;background:{color};"></div>'
            f'<span>{value}</span></div>'
        )
    return "\n".join(parts)


def legend_html() -> str:
    items = []
    for mode, color in MODE_COLORS_SOLID.items():
        items.append(
            f'<span class="legend-item"><span class="legend-swatch" style="background:{color};"></span>{html.escape(mode)}</span>'
        )
    items.append(
        '<span class="legend-item"><span class="legend-swatch" style="background:#fff;border:2px solid #c0392b;'
        'color:#c0392b;font-size:10px;text-align:center;line-height:12px;">~</span>Élongation</span>'
    )
    return '<div class="legend">' + ' '.join(items) + '</div>'


def display_panel_html() -> str:
    """Section « Affichage » — contrôle visuel uniquement."""
    # Modes d'expression
    mode_cbs = []
    for mode in MODES:
        mode_cbs.append(
            f'<label class="mode-label" style="color:{MODE_COLORS_SOLID[mode]}">'
            f'<input type="checkbox" class="display-mode" value="{html.escape(mode, quote=True)}" checked> '
            f'{html.escape(mode)}</label>'
        )
    # Marqueurs linguistiques (binaires + élongations)
    bin_cbs = []
    for bf in BINARY_FEATURES:
        bin_cbs.append(
            f'<label><input type="checkbox" class="display-binary" data-feature="{bf["attr"]}"> '
            f'{html.escape(bf["label"])}</label>'
        )
    # Élongations — même groupe
    elong_cb = (
        '<label style="color:#c0392b;font-weight:600">'
        '<input type="checkbox" id="display-elongations" checked> Élongations</label>'
    )
    # Annotations qualitatives
    qual_cbs = []
    for qf in QUAL_FEATURES:
        qual_cbs.append(
            f'<label><input type="checkbox" class="display-qual" data-feature="{qf["attr"]}"> '
            f'{html.escape(qf["label"])}</label>'
        )

    return (
        '<div class="panel display-panel open">'
        '<div class="panel-header"><span class="panel-title">🎨 Affichage</span><span class="panel-chevron">▼</span></div>'
        '<div class="panel-body">'
        # Corpus name toggle
        '<div class="ctrl-group"><div class="ctrl-group-title">Général</div>'
        '<div class="ctrl-row"><label><input type="checkbox" id="display-corpus-names"> '
        'Noms de corpus</label></div></div>'
        # Modes
        '<div class="ctrl-group"><div class="ctrl-group-title">Modes d\'expression</div>'
        f'<div class="ctrl-row">{"".join(mode_cbs)}</div></div>'
        # Marqueurs linguistiques (binaires + élongations)
        '<div class="ctrl-group"><div class="ctrl-group-title">Marqueurs linguistiques</div>'
        f'<div class="ctrl-row">{"".join(bin_cbs)}{elong_cb}</div></div>'
        # Qualitatives
        '<div class="ctrl-group"><div class="ctrl-group-title">Annotations qualitatives</div>'
        f'<div class="ctrl-row">{"".join(qual_cbs)}</div></div>'
        '</div></div>'
    )


def filter_panel_html(qual_values: dict[str, list[str]], total: int) -> str:
    """Section « Filtrer » — sélection des données."""
    # Corpus
    corpus_cbs = []
    for ds in DATASETS:
        corpus_cbs.append(
            f'<label><input type="checkbox" class="filter-corpus" value="{ds["key"]}" checked> '
            f'{html.escape(ds["label"])}</label>'
        )
    # Type de texte
    type_cbs = (
        '<label><input type="checkbox" id="filter-with-emo" checked> Avec émotions</label>'
        '<label><input type="checkbox" id="filter-without-emo" checked> Sans émotions</label>'
        '<label><input type="checkbox" id="filter-with-elong"> Avec élongation uniquement</label>'
    )
    # Mode d'expression
    mode_cbs = []
    for mode in MODES:
        mode_cbs.append(
            f'<label class="mode-label" style="color:{MODE_COLORS_SOLID[mode]}">'
            f'<input type="checkbox" class="filter-mode" value="{html.escape(mode, quote=True)}"> '
            f'{html.escape(mode)}</label>'
        )
    # Binaires (+ élongation)
    bin_cbs = []
    for bf in BINARY_FEATURES:
        bin_cbs.append(
            f'<label><input type="checkbox" class="filter-binary" data-feature="{bf["attr"]}"> '
            f'{html.escape(bf["label"])}</label>'
        )
    bin_cbs.append(
        '<label style="color:#c0392b;font-weight:600">'
        '<input type="checkbox" class="filter-binary" data-feature="elongation"> Élongation</label>'
    )
    # Qualitatives — groupes dépliables
    qual_groups = []
    for qf in QUAL_FEATURES:
        values = qual_values.get(qf["col"], [])
        if not values:
            continue
        opts = []
        for v in values:
            opts.append(
                f'<label><input type="checkbox" class="filter-qual" data-feature="{qf["attr"]}" '
                f'value="{html.escape(v, quote=True)}"> {html.escape(v)}</label>'
            )
        qual_groups.append(
            f'<details class="qual-filter-group"><summary>{html.escape(qf["label"])}</summary>'
            f'<div class="qual-opts">{"".join(opts)}</div></details>'
        )

    return (
        '<div class="panel filter-panel open">'
        '<div class="panel-header"><span class="panel-title">🔍 Filtrer</span><span class="panel-chevron">▼</span></div>'
        '<div class="panel-body">'
        # Corpus
        '<div class="ctrl-group"><div class="ctrl-group-title">Corpus</div>'
        f'<div class="ctrl-row">{"".join(corpus_cbs)}</div></div>'
        # Type
        '<div class="ctrl-group"><div class="ctrl-group-title">Type de texte</div>'
        f'<div class="ctrl-row">{type_cbs}</div></div>'
        # Mode d'expression
        '<div class="ctrl-group"><div class="ctrl-group-title">Mode d\'expression</div>'
        f'<div class="ctrl-row">{"".join(mode_cbs)}</div></div>'
        # Binaires
        '<div class="ctrl-group"><div class="ctrl-group-title">Traits binaires (ET logique)</div>'
        f'<div class="ctrl-row">{"".join(bin_cbs)}</div></div>'
        # Qualitatives
        '<div class="ctrl-group"><div class="ctrl-group-title">Variables qualitatives</div>'
        f'<div class="ctrl-row" style="flex-direction:column;align-items:stretch;gap:2px">{"".join(qual_groups)}</div></div>'
        # Recherche
        '<div class="search-row">'
        '<input type="text" id="search-box" placeholder="Rechercher un mot dans les textes…">'
        f'<span class="counter">Textes affichés : <span id="shown-count">{total}</span> / {total}</span>'
        '</div>'
        '</div></div>'
    )


def render_record(dataset: dict, row: dict) -> str:
    """Rendu d'un enregistrement comme carte HTML avec data-* attributes."""
    text = str(row.get("text") or row.get("TEXT") or "")
    active_modes = get_active_modes(row)
    active_emotions = get_active_emotions(row)
    has_emo = "1" if as_int(row.get("Emo")) == 1 else "0"

    # Badges émotions (toujours visibles)
    meta_parts = active_emotions
    meta_html = ''.join(
        f'<span class="doc-meta">{html.escape(str(p))}</span>' for p in meta_parts
    )

    # Badges binaires (visibilité contrôlée par CSS body classes)
    bin_badges = ""
    for bf in BINARY_FEATURES:
        if as_int(row.get(bf["col"])) == 1:
            bin_badges += f'<span class="badge badge-{bf["attr"]}">{html.escape(bf["label"])}</span>'

    # Badges qualitatives
    qual_badges = ""
    for qf in QUAL_FEATURES:
        v = row.get(qf["col"], "")
        if v:
            qual_badges += f'<span class="badge-qual badge-{qf["attr"]}">{html.escape(qf["label"])}: {html.escape(str(v))}</span>'

    # Élongation spans
    elong_spans: list[tuple[int, int]] = []
    elong_json = row.get("elongation_spans")
    if elong_json:
        try:
            parsed = json.loads(elong_json)
            elong_spans = [(item["start"], item["end"]) for item in parsed if "start" in item and "end" in item]
        except (json.JSONDecodeError, TypeError):
            pass
    has_elongation = "1" if elong_spans else "0"

    no_emo_class = " no-emo" if has_emo == "0" else ""

    # data-* attributes
    data_attrs = (
        f'data-corpus="{dataset["key"]}" '
        f'data-modes="{html.escape(",".join(active_modes), quote=True)}" '
        f'data-emos="{html.escape(",".join(active_emotions), quote=True)}" '
        f'data-has-emo="{has_emo}" '
        f'data-has-elongation="{has_elongation}"'
    )
    for bf in BINARY_FEATURES:
        data_attrs += f' data-{bf["attr"]}="{as_int(row.get(bf["col"]))}"'
    for qf in QUAL_FEATURES:
        v = row.get(qf["col"], "")
        data_attrs += f' data-{qf["attr"]}="{html.escape(str(v), quote=True)}"'

    return (
        f'<div class="doc-container{no_emo_class}" {data_attrs}>'
        '<div class="doc-header">'
        f'<span class="doc-corpus-badge">{html.escape(dataset["label"])}</span>'
        f'{meta_html}{bin_badges}{qual_badges}'
        '</div>'
        f'<div class="doc-text">{highlight_text_with_spans(text, row.get("spans_json"), elong_spans)}</div>'
        '</div>'
    )


def render_dataset_section(dataset: dict) -> str:
    parts = [f'<section class="corpus-section" data-corpus="{dataset["key"]}">']
    for record in dataset["records"]:
        parts.append(render_record(dataset, record))
    parts.append('</section>')
    return '\n'.join(parts)


# ── Assemblage final ──────────────────────────────────────────────────────────

def generate_combined_html(datasets: list[dict], out_path: Path) -> None:
    all_records = [r for ds in datasets for r in ds["records"]]
    overall_stats = compute_stats(all_records)
    qual_values = collect_qual_values(all_records)

    html_parts = [
        '<!DOCTYPE html>',
        '<html lang="fr">',
        '<head>',
        '<meta charset="utf-8"/>',
        '<title>Visualisation des annotations sur le corpus de CyberHarcèlement</title>',
        '<link rel="preconnect" href="https://fonts.googleapis.com">',
        '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>',
        '<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">',
        f'<style>{CSS}</style>',
        '</head>',
        '<body class="show-elongations">',
        # Header
        '<div class="app-header">',
        '<h1>Visualisation des annotations sur le corpus de CyberHarcèlement</h1>',
        '</div>',
        '<div class="main-content">',
        # Top controls
        '<div class="top-controls">',
        '<button id="btn-toggle-stats" class="btn">Afficher les statistiques</button>',
        '</div>',
        # Stats
        '<div id="stats-section">',
        '<div class="stats-grid">',
        '<div class="stat-box"><h3>Distribution globale des modes</h3>',
        bar_chart_html(overall_stats["mode_counts"], MODE_COLORS_SOLID),
        '</div>',
        '<div class="stat-box"><h3>Distribution globale des émotions</h3>',
        bar_chart_html(overall_stats["emo_counts"]),
        '</div>',
        '</div></div>',
        # Legend
        legend_html(),
        # Panels
        display_panel_html(),
        filter_panel_html(qual_values, overall_stats["total"]),
    ]
    # Corpus sections
    for ds in datasets:
        html_parts.append(render_dataset_section(ds))
    # Footer / script
    html_parts.append('</div>')  # .main-content
    html_parts.append(f'<script>{JAVASCRIPT}</script>')
    html_parts.append('</body></html>')

    out_path.write_text('\n'.join(html_parts), encoding='utf-8')


# ── Chargement des données ────────────────────────────────────────────────────

def load_dataset(base_dir: Path, spec: dict) -> dict:
    path = base_dir / spec["path"]
    headers, rows = read_xlsx(path)
    records = rows_to_dicts(headers, rows)
    return {
        "key": spec["key"],
        "label": spec["label"],
        "path": path,
        "records": records,
        "stats": compute_stats(records),
    }


def main() -> None:
    datasets = []
    missing = []
    for spec in DATASETS:
        path = BASE_DIR / spec["path"]
        if path.exists():
            print(f"Processing {spec['label']}…")
            datasets.append(load_dataset(BASE_DIR, spec))
        else:
            missing.append(path)

    if missing:
        raise FileNotFoundError(f"Missing input files:\n" + "\n".join(f"- {p}" for p in missing))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    generate_combined_html(datasets, OUTPUT_PATH)
    total = sum(ds["stats"]["total"] for ds in datasets)
    print(f"Generated {OUTPUT_PATH} ({total} texts across {len(datasets)} corpora).")


if __name__ == "__main__":
    main()
