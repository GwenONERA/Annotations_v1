#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Add an ``elongation_spans`` column to every XLSX corpus file.

The column contains a **JSON array** of span objects, each with:
  - ``word``  : the elongated word/segment as it appears in TEXT
  - ``start`` : character offset (0-indexed) of the first character
  - ``end``   : character offset (exclusive) of the last character

Example value:
  [{"word": "mdrrrr", "start": 0, "end": 6}, {"word": "grvvv", "start": 7, "end": 12}]

Elongation sources (merged & deduplicated):
  1. **Auto-detection** — any run of 3+ identical characters
  2. **JSONL annotations** — entries with verdict == "elongation"
  3. **Manual additions** — user-identified segments with < 3 repeats

Run once to populate the column, then ``generate_html_viz.py`` reads
directly from it — no JSONL loading at runtime.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

# ── Configuration ────────────────────────────────────────────────────────────

BASE_DIR = Path("/home/gwen/Annotations_v1")

XLSX_FILES = [
    BASE_DIR / "outputs" / "homophobie_annotations_gold_flat_updated.xlsx",
    BASE_DIR / "outputs" / "obésité_annotations_gold_flat_updated.xlsx",
    BASE_DIR / "outputs" / "racisme_annotations_gold_flat_updated.xlsx",
    BASE_DIR / "outputs" / "religion_annotations_gold_flat_updated.xlsx",
]

ELONGATION_JSONL = BASE_DIR / "experimentations" / "elongations" / "elongations_annotated.jsonl"

SPAN_COL_NAME = "elongation_spans"

# Manual elongation segments that have < 3 identical consecutive characters
# and therefore escape the auto-detector.  These are case-insensitive.
MANUAL_SEGMENTS: list[str] = [
    "eouhh",   # 2× 'h'  — "le muslim 2.0 se reveill eouhh attention !!"
]


# ── Auto-detection ───────────────────────────────────────────────────────────

def detect_elongation_spans(text: str) -> list[dict]:
    """Find words/segments containing 3+ consecutive identical characters.

    For alphabetic runs the span is expanded to word boundaries;
    for punctuation runs (``???``, ``...``) the span covers the run itself.
    """
    if not text:
        return []

    spans: list[dict] = []
    seen_ranges: set[tuple[int, int]] = set()

    for m in re.finditer(r"(.)\1{2,}", text):
        char = m.group(1)
        run_start, run_end = m.start(), m.end()

        if char.isalpha():
            # expand to word boundaries
            word_start = run_start
            while word_start > 0 and not text[word_start - 1].isspace():
                word_start -= 1
            word_end = run_end
            while word_end < len(text) and not text[word_end].isspace():
                word_end += 1
        else:
            word_start, word_end = run_start, run_end

        key = (word_start, word_end)
        if key in seen_ranges:
            continue
        seen_ranges.add(key)
        spans.append({"word": text[word_start:word_end], "start": word_start, "end": word_end})

    return spans


# ── Manual additions ─────────────────────────────────────────────────────────

def find_manual_spans(text: str) -> list[dict]:
    """Locate manually-listed segments (case-insensitive) in *text*."""
    spans: list[dict] = []
    text_lower = text.lower()
    for seg in MANUAL_SEGMENTS:
        idx = text_lower.find(seg.lower())
        if idx < 0:
            continue
        # Expand to word boundaries
        ws = idx
        while ws > 0 and not text[ws - 1].isspace():
            ws -= 1
        we = idx + len(seg)
        while we < len(text) and not text[we].isspace():
            we += 1
        spans.append({"word": text[ws:we], "start": ws, "end": we})
    return spans


# ── JSONL loading ────────────────────────────────────────────────────────────

def load_jsonl_elongations(jsonl_path: Path) -> dict[str, list[dict]]:
    """Return ``{texte_brut: [{raw_word, pos_hint}, …]}`` for verdict == elongation."""
    result: dict[str, list[dict]] = {}
    if not jsonl_path.exists():
        print(f"  Warning: JSONL not found: {jsonl_path}")
        return result

    with open(jsonl_path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            entry = json.loads(line)
            if entry.get("verdict") != "elongation":
                continue

            texte = entry.get("texte_brut", "")
            detail = entry.get("detail_elongation", "")
            m = re.search(r"\*\*Mot brut\*\*\s*:\s*`([^`]*)`", detail)
            if not m:
                continue
            raw_word = m.group(1)

            # position hint from record_id
            record_id = entry.get("record_id", "")
            parquet_id = entry.get("parquet_id", "")
            pos_hint = -1
            prefix = str(parquet_id) + "_"
            if record_id.startswith(prefix):
                remainder = record_id[len(prefix):]
                try:
                    pos_hint = int(remainder.split("_")[0])
                except ValueError:
                    pass

            result.setdefault(texte, []).append({"raw_word": raw_word, "pos_hint": pos_hint})

    return result


def jsonl_spans_for_text(text: str, entries: list[dict]) -> list[dict]:
    """Convert JSONL entries into span dicts for a given *text*."""
    spans: list[dict] = []
    text_lower = text.lower()
    used: set[int] = set()

    for entry in entries:
        raw_word = entry["raw_word"]
        pos_hint = entry["pos_hint"]
        word_lower = raw_word.lower()

        # all occurrences
        candidates: list[int] = []
        s = 0
        while True:
            idx = text_lower.find(word_lower, s)
            if idx < 0:
                break
            candidates.append(idx)
            s = idx + 1

        if not candidates:
            continue

        best = None
        for c in candidates:
            if c in used:
                continue
            if pos_hint >= 0 and c <= pos_hint < c + len(raw_word):
                best = c
                break
            if best is None:
                best = c
        if best is None:
            best = candidates[0]

        used.add(best)
        spans.append({"word": raw_word, "start": best, "end": best + len(raw_word)})

    return spans


# ── Span merging ─────────────────────────────────────────────────────────────

def merge_spans(all_spans: list[dict], text: str) -> list[dict]:
    """Merge overlapping/adjacent spans.  Returns sorted, non-overlapping list."""
    if not all_spans:
        return []

    sorted_spans = sorted(all_spans, key=lambda s: (s["start"], -s["end"]))
    merged = [{"start": sorted_spans[0]["start"], "end": sorted_spans[0]["end"]}]

    for s in sorted_spans[1:]:
        last = merged[-1]
        if s["start"] <= last["end"]:
            last["end"] = max(last["end"], s["end"])
        else:
            merged.append({"start": s["start"], "end": s["end"]})

    # Reconstruct word field from the text using the merged range
    for span in merged:
        span["word"] = text[span["start"]:span["end"]]

    return merged


# ── Process one XLSX ─────────────────────────────────────────────────────────

def process_xlsx(xlsx_path: Path, jsonl_data: dict[str, list[dict]]) -> int:
    """Add/update *elongation_spans* column.  Returns number of rows with spans."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    text_col = headers.index("TEXT") + 1  # 1-indexed

    # Find or create the target column
    if SPAN_COL_NAME in headers:
        span_col = headers.index(SPAN_COL_NAME) + 1
    else:
        span_col = len(headers) + 1
        ws.cell(row=1, column=span_col, value=SPAN_COL_NAME)

    count = 0
    for row_idx in range(2, ws.max_row + 1):
        text_val = ws.cell(row=row_idx, column=text_col).value
        text = str(text_val) if text_val is not None else ""
        if not text or text == "None":
            ws.cell(row=row_idx, column=span_col, value="")
            continue

        # Gather spans from all three sources
        all_spans: list[dict] = []
        all_spans.extend(detect_elongation_spans(text))
        if text in jsonl_data:
            all_spans.extend(jsonl_spans_for_text(text, jsonl_data[text]))
        all_spans.extend(find_manual_spans(text))

        merged = merge_spans(all_spans, text)

        if merged:
            ws.cell(row=row_idx, column=span_col, value=json.dumps(merged, ensure_ascii=False))
            count += 1
        else:
            ws.cell(row=row_idx, column=span_col, value="")

    wb.save(xlsx_path)
    wb.close()
    return count


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Loading JSONL elongation annotations…")
    jsonl_data = load_jsonl_elongations(ELONGATION_JSONL)
    n_entries = sum(len(v) for v in jsonl_data.values())
    print(f"  {n_entries} entries across {len(jsonl_data)} unique texts")

    for xlsx_path in XLSX_FILES:
        if not xlsx_path.exists():
            print(f"\n  SKIP (not found): {xlsx_path.name}")
            continue
        print(f"\nProcessing {xlsx_path.name}…")
        count = process_xlsx(xlsx_path, jsonl_data)
        print(f"  → {count} rows now have elongation_spans")

    print("\n✓ Done. All XLSX files updated with '{SPAN_COL_NAME}' column.")


if __name__ == "__main__":
    main()
